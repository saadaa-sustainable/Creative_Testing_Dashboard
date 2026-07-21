"""Build a structured changelog Excel from git history since GUIDEBOOK.docx
was last regenerated (2026-07-02) up to the most recent commit + uncommitted
work-in-progress.

Sheets:
  1) Overview               — how to read the log
  2) Changes                — one row per change, with area/type/impact
  3) New DB tables/views    — Postgres additions since guidebook
  4) Uncommitted (WIP)      — my session work from 2026-07-20 not yet in git
"""
import os, sys, subprocess, re
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

REPO = Path("D:/Creative_Testing_Dashboard")
OUT  = REPO / "CHANGELOG_since_2026-07-02.xlsx"

def git_log():
    """Return list of (sha, iso_date, subject) for commits after 2026-07-02."""
    out = subprocess.check_output(
        ["git", "-C", str(REPO), "log", "--format=%h|%ad|%s", "--date=short",
         "--since=2026-07-02"],
        text=True, encoding="utf-8", errors="replace")
    rows = []
    for line in out.strip().splitlines():
        parts = line.split("|", 2)
        if len(parts) == 3:
            rows.append(tuple(parts))
    rows.reverse()   # chronological
    return rows

def git_files(sha):
    out = subprocess.check_output(
        ["git", "-C", str(REPO), "show", "--name-only", "--pretty=", sha],
        text=True, encoding="utf-8", errors="replace")
    return [l.strip() for l in out.splitlines() if l.strip()]

# ── Classifier ─────────────────────────────────────────────
# Uses commit subject + touched files to bucket into Area (Backend / Frontend /
# Pipeline / Attribution / Schema / Docs) and Type (feature/fix/refactor/data).
def classify(subject, files):
    subj = subject.lower()
    fs = " ".join(files).lower()

    # Area
    if "guidebook" in fs or subj.startswith("regen guidebook"):
        area = "Docs"
    elif "index_v2.html" in fs or "assets/" in fs or "dashboard.js" in fs:
        area = "Frontend"
    elif "attribution" in subj or "rebuild_attribution_orders" in fs or "reattribute" in fs:
        area = "Attribution"
    elif ("primary_sync" in fs or "refresh_" in fs or "pipeline" in subj
          or "fetch_" in fs or "sync" in fs or "backfill" in subj):
        area = "Backend / Pipeline"
    elif "migration" in fs or "schema" in fs or ".sql" in fs:
        area = "DB Schema"
    else:
        area = "Backend / Pipeline"

    # Type
    if any(k in subj for k in ["fix", "revert", "regression", "silent"]):
        typ = "Bug fix"
    elif any(k in subj for k in ["add ", "new ", "introduce", "scaffold", "wire",
                                 "onboard", "expose"]):
        typ = "New feature"
    elif any(k in subj for k in ["refactor", "split", "rename", "cleanup",
                                 "reorg", "promote", "revamp", "rebuild"]):
        typ = "Refactor"
    elif any(k in subj for k in ["cache-bust", "cache bust", "responsive",
                                 "polish", "style", "widen", "narrow",
                                 "column picker", "toggle", "reflow", "ux"]):
        typ = "UX polish"
    elif any(k in subj for k in ["pipeline", "sync", "backfill", "importer",
                                 "sheet importer", "nightly", "daily", "hook",
                                 "matview"]):
        typ = "Data pipeline"
    else:
        typ = "Enhancement"

    # Impact — coarse heuristic on subject keywords
    if any(k in subj for k in ["new_incr_table", "google ads", "primary_sync",
                               "attribution table", "scope-first", "5-step"]):
        impact = "High"
    elif any(k in subj for k in ["cache-bust", "widen", "narrow", "chip",
                                 "colour", "color", "info button", "banner",
                                 "font size", "responsive polish"]):
        impact = "Low"
    else:
        impact = "Medium"

    return area, typ, impact

# ── DB additions since guidebook ─────────────────────────
# Hand-curated based on the commit history + this session's work.
# Kept as a static table so the Excel doesn't need live DB access to generate.
DB_ADDITIONS = [
    # (name, kind, since_commit, purpose)
    ("google_ads_primary",        "table", "d2bfa98 (2026-07-16)",
     "Meta-style daily grain for Google Ads insights (ad×date)"),
    ("google_ads_summary",        "table", "d2bfa98 (2026-07-16)",
     "Lifetime rollup of Google Ads per ad_id — feeds AE table's Google view"),
    ("google_campaigns_meta",     "table", "d2bfa98 (2026-07-16)",
     "One row per campaign incl. PMax/YouTube (metadata only, no metrics)"),
    ("ig_media",                  "table", "d2bfa98 (2026-07-16)",
     "Instagram media metadata + insights for 3 IG accounts"),
    ("ad_asset_ids",              "table", "aba109e (2026-07-13)",
     "Manual ad_id → asset_id mapping; editable inline in dashboard"),
    ("ae_reach_recent",           "table", "59a8fd0 (2026-07-13)",
     "Dedup'd primary+backfill reach snapshot for Incremental Reach view"),
    ("new_incr_table",            "table", "49abef6 (2026-07-14)",
     "Ad-level daily cum/incr reach + spend series (rebuilt every sync)"),
    ("primary_adset_table",       "table", "SESSION 2026-07-20 (uncommitted)",
     "Meta /insights level=adset daily grain — deduped reach (37 mo backfilled)"),
    ("primary_camp_table",        "table", "SESSION 2026-07-20 (uncommitted)",
     "Meta /insights level=campaign daily grain — deduped reach (37 mo)"),
    ("new_incr_adset_table",      "table", "SESSION 2026-07-20 (uncommitted)",
     "Adset-level cum/incr reach series (mirrors new_incr_table shape)"),
    ("new_incr_camp_table",       "table", "SESSION 2026-07-20 (uncommitted)",
     "Campaign-level cum/incr reach series"),
    ("rebuild_new_incr_table",    "function", "49abef6 (2026-07-14)",
     "Postgres fn — rebuilds public.new_incr_table from primary+backfill"),
    ("rebuild_new_incr_adset_table","function","SESSION 2026-07-20 (uncommitted)",
     "Postgres fn — rebuilds adset-level new_incr from primary_adset_table"),
    ("rebuild_new_incr_camp_table","function", "SESSION 2026-07-20 (uncommitted)",
     "Postgres fn — rebuilds campaign-level new_incr"),
    ("get_reach_by_window",       "function", "2a2f92c (2026-07-10)",
     "RPC returning per-ad reach_first/reach_last/reach_incr/reach_sum/spend_sum"),
    ("shopify_ad_attribution.customer_id",      "column", "49abef6 (2026-07-14)",
     "Customer ID column added — backfilled via backfill_customer_info.py"),
    ("shopify_ad_attribution.customer_num_orders","column","49abef6 (2026-07-14)",
     "Lifetime order count per customer at attribution time"),
    ("shopify_ad_attribution.contact_email",    "column", "49abef6 (2026-07-14)",
     "Buyer email — hover on Customer ID cell to see"),
    ("google_ads_summary.campaign_status",      "column", "SESSION 2026-07-20 (uncommitted)",
     "Campaign status — was defaulting to ENABLED for everything"),
    ("google_ads_summary.ad_group_status",      "column", "SESSION 2026-07-20 (uncommitted)",
     "Ad group status — same bug as above"),
]

# ── Uncommitted session work (2026-07-20) ────────────────
SESSION_WIP = [
    # (file, area, change, impact)
    ("backend/primary_sync.py", "Backend / Pipeline",
     "Fix silent placeholder failure: /ads endpoint limit 500 → 50 (Fourth Ad "
     "Account hit Meta 500 on nested creative fields). Also added Meta 400 "
     "rate-limit retry logic in _get.", "High"),
    ("backend/primary_sync.py", "Backend / Pipeline",
     "Wire link_url_asset breakdown overlay in sync(): one call per account "
     "per sync, Meta's delivered URL wins over the 6-path creative walk when "
     "present. Captures URL-rotation for ~260 Advantage+ URL-testing ads.", "Medium"),
    ("backend/rebuild_attribution_orders.py", "Attribution",
     "_scoped_match token-subset step: utm tokens ⊆ ad_name tokens, single "
     "hit wins. Catches 'SMFLK_IFAD-vinitsikariya_16/12/25' → correct ad_id "
     "when older substring checks failed.", "High"),
    ("backend/rebuild_attribution_orders.py", "Attribution",
     "_scoped_match ratio-tiebreak: on multi-hit, rank ads by MAX(|utm∩ad|/|"
     "ad_tokens|), pick winner if ratio ≥0.6 AND margin ≥0.15. Correctly "
     "resolves SDVPL case where 3 adset siblings all subset-match.", "High"),
    ("backend/refresh_new_incr_table.py", "Backend / Pipeline",
     "Extended to rebuild all 3 new_incr_* tables (ad + adset + campaign) "
     "sequentially instead of just ad-level.", "Medium"),
    ("backend/fetch_adset_camp_reach.py", "Backend / Pipeline",
     "NEW — mirrors primary_sync flow but at level=adset and level=campaign. "
     "37-month backfill: 82,700 adset-day + 18,197 camp-day rows.", "High"),
    ("backend/fetch_archived_ads.py", "Backend / Pipeline",
     "NEW — pulls ARCHIVED ads via filtering=[effective_status IN ['ARCHIVED']] "
     "and lands placeholder rows in backfill_table so attribution engine's "
     "maps include them. 2,411 archived ads captured.", "High"),
    ("backend/reattribute_all.py", "Attribution",
     "Full re-attribute post-archived-fetch: 54,244 UPDATEs applied (orders "
     "that pointed at archived ads now resolve to those ads).", "High"),
    ("assets/dashboard.js + index_v2.html", "Frontend",
     "COMMITTED d5d0776: prefix-match Google tier labels in aiStep + "
     "aiGoogleStep so G1-G4 render as matched (DB stores G1_ad_id / "
     "G4_campaign_only variants). Cache-buster bumped to gtier_matched3.", "High"),
]

# ── Build workbook ─────────────────────────────────────
wb = Workbook()

# Style helpers
hdr_font  = Font(bold=True, color="FFFFFF", size=11)
hdr_fill  = PatternFill("solid", fgColor="2C3E50")
sub_fill  = PatternFill("solid", fgColor="ECF0F1")
warn_fill = PatternFill("solid", fgColor="FFF3CD")
box       = Border(left=Side(style="thin", color="BDC3C7"),
                   right=Side(style="thin", color="BDC3C7"),
                   top=Side(style="thin", color="BDC3C7"),
                   bottom=Side(style="thin", color="BDC3C7"))
wrap = Alignment(wrap_text=True, vertical="top", horizontal="left")

def style_header_row(ws, row=1):
    for cell in ws[row]:
        cell.font = hdr_font
        cell.fill = hdr_fill
        cell.alignment = Alignment(horizontal="left", vertical="center")

def apply_body(ws, start_row=2):
    for row in ws.iter_rows(min_row=start_row):
        for cell in row:
            cell.alignment = wrap
            cell.border = box

def set_widths(ws, widths):
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w

# ── Sheet 1: Overview ─────────────────────
ws = wb.active
ws.title = "Overview"
ws["A1"] = "Saadaa Creative Testing Dashboard — Changelog"
ws["A1"].font = Font(bold=True, size=14)
ws["A2"] = "Baseline: GUIDEBOOK.docx as of 2026-07-02 (commit 2194c02)"
ws["A3"] = "Cutoff:   HEAD as of 2026-07-20  +  uncommitted session work"
ws["A4"] = ""
ws["A5"] = "Sheets in this workbook"
ws["A5"].font = Font(bold=True)
overview = [
    ("Changes",              "One row per commit or WIP change with area/type/impact tags"),
    ("New DB Tables",        "Postgres tables/views/columns/functions added since baseline"),
    ("Uncommitted (WIP)",    "Session work from 2026-07-20 not yet in git (details)"),
]
for i, (name, desc) in enumerate(overview, start=6):
    ws.cell(row=i, column=1, value=name).font = Font(bold=True)
    ws.cell(row=i, column=2, value=desc)
ws["A10"] = ""
ws["A11"] = "Legend"
ws["A11"].font = Font(bold=True)
legend = [
    ("Area",    "Frontend | Backend/Pipeline | Attribution | DB Schema | Docs"),
    ("Type",    "New feature | Bug fix | Refactor | UX polish | Data pipeline | Enhancement"),
    ("Impact",  "High = user-visible or data-shape change · Medium = component-level · Low = cosmetic"),
]
for i, (k, v) in enumerate(legend, start=12):
    ws.cell(row=i, column=1, value=k).font = Font(bold=True)
    ws.cell(row=i, column=2, value=v)
set_widths(ws, [22, 90])

# ── Sheet 2: Changes ─────────────────────
ws = wb.create_sheet("Changes")
headers = ["#", "Date", "Commit", "Area", "Type", "Impact", "Summary"]
ws.append(headers)
style_header_row(ws)

commits = git_log()
for i, (sha, dt, subj) in enumerate(commits, start=1):
    files = git_files(sha)
    area, typ, impact = classify(subj, files)
    ws.append([i, dt, sha, area, typ, impact, subj])
apply_body(ws)
set_widths(ws, [5, 12, 10, 22, 16, 10, 90])
ws.freeze_panes = "A2"
ws.auto_filter.ref = ws.dimensions

# ── Sheet 3: New DB Tables ─────────────────────
ws = wb.create_sheet("New DB Tables")
headers = ["#", "Object", "Kind", "Added in", "Purpose"]
ws.append(headers)
style_header_row(ws)
for i, row in enumerate(DB_ADDITIONS, start=1):
    ws.append([i] + list(row))
apply_body(ws)
set_widths(ws, [5, 40, 12, 35, 70])
ws.freeze_panes = "A2"

# ── Sheet 4: Uncommitted (WIP) ─────────────────────
ws = wb.create_sheet("Uncommitted (WIP)")
headers = ["#", "File / Component", "Area", "Change", "Impact"]
ws.append(headers)
style_header_row(ws)
for i, row in enumerate(SESSION_WIP, start=1):
    ws.append([i] + list(row))
apply_body(ws)
# Highlight uncommitted sheet with warning fill on the header
for cell in ws[1]:
    cell.fill = hdr_fill  # keep header dark
# Set widths
set_widths(ws, [5, 34, 22, 90, 10])
ws.freeze_panes = "A2"

wb.save(OUT)
print(f"wrote {len(commits)} commits + {len(DB_ADDITIONS)} DB objects + {len(SESSION_WIP)} WIP items")
print(f"→ {OUT}")
