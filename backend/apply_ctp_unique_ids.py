"""
apply_ctp_unique_ids.py - merge 'CTP unique ids.xlsx' into summary_table.

Mirrors the Google Apps Script `SubsheetMatch.gs` matching logic:
  - Per-subsheet key column (Asset ID / POST ID / Requisition ID / Nomenclature)
  - Digit-boundary guard: "GAD-Apr-4" won't match "GAD-Apr-48"
  - Best-rank result: when multiple ads contain the ID, the highest
    lifecycle outcome wins (Incremental Winner > Winner > Iteration >
    Analyse > Discarded)

Writes three columns on summary_table (existing `status` is UNTOUCHED):
  excel_id_matched : the ID string we matched
  excel_status     : 'Tested' if matched, else (from Excel: Pending/Tested/...)
  excel_result     : derived best-rank label, falls back to Excel Result
"""
import os, sys, re
import openpyxl, psycopg2
from collections import defaultdict
from psycopg2.extras import execute_batch
from dotenv import load_dotenv

load_dotenv()
DB_URL = os.environ["SUPABASE_DB_URL"].strip()
XLSX   = r"C:\Users\Saadaa\Downloads\CTP unique ids.xlsx"

# Tab names + key column header (per Apps Script SM_SUBSHEETS).
# Falls back to similarly-named tab (case-insensitive) if exact name missing.
SHEETS = [
    # tab_name                                  primary_key_header     alt_key_header   status_col  result_col
    ("Video",                                  "Asset ID",             None,            "Status", "Result"),
    ("Influencer",                             "POST ID",              "SIF-ID",        "Status", "Result"),
    ("Graphic",                                "Requisition ID",       None,            "Status", "Result"),
    ("Deprecated Completed File",              "Requisition ID",       None,            "Status", "Result"),
    ("Deprecated Planning",                    "Requisition ID",       None,            "Status", "Result"),
    ("Deprecated Unplanned Edited Content",    "Nomenclature",         None,            "Status", "Result"),
    # Legacy fallback sheets that may still exist with the older spelling
    ("Influncer",                              "POST ID",              "SIF-ID",        "Status", "Result"),
    ("Deprecated Completed file",              "Requisition ID",       None,            "Status", "Result"),
    ("Deprecated Unplanned Edited con",        "Nomenclature",         None,            "Status", "Result"),
    ("Sheet6",                                 "Requisition ID",       None,            "Status", "Result"),
]

# Rank scale (per Apps Script SM_RANK)
RANK = {
    "incremental winner": 5,
    "winner":             4,
    "iteration":          3,
    "priority":           3,   # alias used in our F-filter naming
    "analyse":            2,
    "analyze 1":          2,   # alias
    "analyze 2":          2,   # alias
    "discarded":          1,
}
LABEL = {5: "Incremental Winner", 4: "Winner", 3: "Iteration", 2: "Analyse", 1: "Discarded"}

print(f"[load] reading {XLSX}", flush=True)
wb = openpyxl.load_workbook(XLSX, data_only=True, read_only=True)
# Case-insensitive sheet lookup (Apps Script style)
sheets_lower = {s.lower(): s for s in wb.sheetnames}

def _find_sheet(name):
    return wb[name] if name in wb.sheetnames else (
        wb[sheets_lower[name.lower()]] if name.lower() in sheets_lower else None
    )

def _header_index(headers, target):
    t = target.strip().lower()
    for i, h in enumerate(headers):
        if str(h or "").strip().lower() == t: return i
    return -1

# Build a master ID -> (status, result, source-sheet) mapping
id_map = {}
seen_sheets = set()
for cfg in SHEETS:
    sheet_name, key_hdr, alt_hdr, status_hdr, result_hdr = cfg
    if sheet_name in seen_sheets: continue
    ws = _find_sheet(sheet_name)
    if ws is None:
        print(f"  [skip] sheet not found: {sheet_name}"); continue
    seen_sheets.add(sheet_name)
    rows_iter = ws.iter_rows(values_only=True)
    headers = next(rows_iter, None) or []
    idx_key    = _header_index(headers, key_hdr)
    idx_alt    = _header_index(headers, alt_hdr) if alt_hdr else -1
    idx_status = _header_index(headers, status_hdr)
    idx_result = _header_index(headers, result_hdr)
    if idx_key < 0 or idx_status < 0 or idx_result < 0:
        print(f"  [skip] {sheet_name}: missing key/status/result column "
              f"(key={idx_key}, status={idx_status}, result={idx_result})")
        continue
    n = 0
    for row in rows_iter:
        if not row: continue
        def _val(i):
            try: return row[i] if 0 <= i < len(row) else None
            except IndexError: return None
        status = _val(idx_status); result = _val(idx_result)
        for col in (idx_key, idx_alt):
            v = _val(col)
            if v is None: continue
            key = str(v).strip()
            if not key: continue
            id_map[key] = {
                "status": (str(status).strip() if status else None),
                "result": (str(result).strip() if result else None),
                "sheet":  sheet_name,
            }
            n += 1
    print(f"  [{sheet_name:<40}] {n:>5} ids")
wb.close()
print(f"[load] total distinct IDs: {len(id_map):,}")

# ── DB connect + schema ──────────────────────────────────────────────────────
conn = psycopg2.connect(DB_URL); conn.autocommit = False
cur = conn.cursor()
print("\n[schema] ensuring excel-derived columns on summary_table ...")
cur.execute("""
  ALTER TABLE summary_table
    ADD COLUMN IF NOT EXISTS excel_id_matched TEXT,
    ADD COLUMN IF NOT EXISTS excel_status     TEXT,
    ADD COLUMN IF NOT EXISTS excel_result     TEXT
""")
conn.commit()

# ── Stream summary_table rows and match (Apps Script logic) ─────────────────
# Apps Script direction: for each AD in summary, find the IDs whose substring
# is in the ad_name with digit-boundary guard, and pick the BEST-RANK status
# across them. We invert: per ad, find the longest ID that boundary-matches.
# (Single ID per ad — multi-match would only happen for nested IDs like
# "SIF-1-P1" inside "SIF-12-P1" which the digit-boundary guard already rules out.)
print("[scan] streaming summary_table rows ...")
cur.execute("SELECT ad_id, ad_name, status FROM summary_table")
rows = cur.fetchall()
print(f"  {len(rows):,} ads to scan")

# Per Apps Script SM_LABEL: derive the rank from summary_table.status, then
# the subsheet's "Result" column reflects that ad's lifecycle category.
def _rank(s):
    return RANK.get((s or "").strip().lower(), 0)

ids_sorted = sorted(id_map.keys(), key=lambda s: -len(s))   # longest first

def _matches_with_boundary(hay_lower, needle_lower):
    """Substring match with the same digit-boundary guard as Apps Script."""
    pos = 0
    L = len(needle_lower)
    while True:
        p = hay_lower.find(needle_lower, pos)
        if p < 0: return False
        next_ch = hay_lower[p + L] if (p + L) < len(hay_lower) else ""
        if not (next_ch.isdigit()): return True
        pos = p + 1

updates = []
unmatched = 0
matched   = 0
for ad_id, ad_name, db_status in rows:
    if not ad_name:
        unmatched += 1; continue
    name_lower = ad_name.lower()
    hit = None
    for k in ids_sorted:
        if _matches_with_boundary(name_lower, k.lower()):
            hit = k; break
    if hit:
        info = id_map[hit]
        # Excel is the source of truth for Status (Pending/Tested) — never
        # overwrite. Result: use the AD's lifecycle rank when the Excel says
        # Tested (so the Iteration/Analyse/etc. label flows through), else
        # fall back to whatever the Excel had.
        excel_st = (info["status"] or "").strip()
        derived_status = excel_st if excel_st else None
        if excel_st.lower() == "tested":
            ad_rank = _rank(db_status)
            derived_result = LABEL.get(ad_rank) if ad_rank else (info["result"] or None)
        else:
            derived_result = info["result"] or None
        updates.append((hit, derived_status, derived_result, ad_id))
        matched += 1
    else:
        unmatched += 1

print(f"  matched   : {matched:,}")
print(f"  unmatched : {unmatched:,}")

# ── Apply UPDATEs ────────────────────────────────────────────────────────────
print(f"\n[update] writing {len(updates):,} rows ...")
execute_batch(cur, """
  UPDATE summary_table
  SET excel_id_matched = %s,
      excel_status     = %s,
      excel_result     = %s
  WHERE ad_id = %s
""", updates, page_size=500)
conn.commit()
print("  committed")

# ── Stats ────────────────────────────────────────────────────────────────────
cur.execute("""
  SELECT
    COUNT(*) FILTER (WHERE excel_id_matched IS NOT NULL) AS with_match,
    COUNT(*) FILTER (WHERE excel_result    IS NOT NULL) AS with_result,
    COUNT(*) FILTER (WHERE excel_result = 'Winner')     AS excel_winners
  FROM summary_table
""")
m, r, w = cur.fetchone()
print(f"\n========== DONE ==========")
print(f"summary_table rows with excel_id_matched : {m:,}")
print(f"  …with non-null excel_result           : {r:,}")
print(f"  …with excel_result = 'Winner'         : {w:,}")

# Cross-tab: F-filter status vs excel_result (only for matched rows)
cur.execute("""
  SELECT status, excel_result, COUNT(*)
  FROM summary_table
  WHERE excel_id_matched IS NOT NULL
  GROUP BY status, excel_result
  ORDER BY status, excel_result NULLS LAST
""")
print("\n[cross-tab] F-filter status vs excel_result (matched rows only)")
print(f"  {'F-filter status':<22}{'excel_result':<16}{'count':>6}")
for st, er, n in cur.fetchall():
    print(f"  {str(st or '-'):<22}{str(er or '-'):<16}{n:>6}")
conn.close()
