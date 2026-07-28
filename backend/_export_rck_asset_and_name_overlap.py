"""For every ad in a target campaign, find matching RCK ads where:
  A) asset_id matches EXACTLY, OR
  B) ad_name is 'kinda the same' — one is a substring of the other after
     stripping common Meta-name prefixes, OR difflib similarity >= 0.75.

Exports to an Excel workbook with ONE tab named after the campaign.
"""
import os, sys, io, re, difflib, psycopg2
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
from dotenv import load_dotenv
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8', errors='backslashreplace')
load_dotenv(override=True)
conn = psycopg2.connect(os.environ['SUPABASE_DB_URL'], connect_timeout=30)

import argparse
_ap = argparse.ArgumentParser()
_ap.add_argument('--campaign', default=None, help='Filter source ads by campaign_name')
_ap.add_argument('--adset',    default=None, help='Filter source ads by adset_name (mutually exclusive with --campaign)')
_ap.add_argument('--workbook', default='RCK_overlap.xlsx',
                 help='Append to this workbook. Sheet is (re)created for the scope.')
_args = _ap.parse_args()
if not _args.campaign and not _args.adset:
    _args.campaign = 'NCP+NCA+BW+HV+INCA+CBO+12/07/26'
if _args.campaign and _args.adset:
    sys.exit('Pass either --campaign OR --adset, not both.')
SCOPE_TYPE = 'campaign' if _args.campaign else 'adset'
CAMP       = _args.campaign or _args.adset   # kept as CAMP for backward-compat with existing code
SCOPE_LABEL = f'{SCOPE_TYPE}: {CAMP}'
SIM_THRESHOLD = 0.75      # difflib similarity cutoff
MIN_CORE_LEN  = 8         # minimum chars for substring match

# ── ad_name normalization: strip common prefixes + placement tokens ────
_PREFIX_TOKENS = [
    'CTP-', 'CLP-', 'ADB_', 'HP-', 'CC+', 'BR_', 'CTP+', 'CLP+',
    'PDP-', 'PDP_', 'CTP-SDAWLP+', 'CTP-SDAWP+', 'CTP-SDWLP+', 'CTP-SDAMK+',
    'CTP-SDVPL+', 'CTP-SDVPK+', 'CTP-SDFLK+', 'CTP-SDFSK+', 'CTP-SDBW+',
    'CTP-SDASK+', 'CTP-SDCP+', 'CTP-SDCSS+', 'CTP-BST+', 'CTP-BST_', 'CTP-W+',
    'CTP-BW+', 'CTP-LE+', 'CTP-FLK+', 'CTP-TW+', 'CTP-FOU+', 'CTP-MSA+',
    'CLP-BST+', 'CLP-SDAWP+', 'CLP-SDCP+', 'CLP-SDFLK+', 'CLP-SDFSK+',
    'CLP-SDPL+', 'CLP-SDWLP+', 'CLP-SDVPL+', 'CLP-MSA+', 'CLP-SMFLK+',
    'CLP-SMCP+', 'SDCP-', 'SMCP_', 'SDAMK_', 'SDAWP_', 'SDAWLP_', 'SDCSS_',
    'SDCSP_', 'SDFLK_', 'SDFSK_', 'SDBW_', 'SDBTJ_', 'SDVPL_', 'SDVPK_',
    'SDASK_', 'SDPL_', 'SDWLP_', 'SDTJ_', 'SDALP_',
]
_SUFFIX_TOKENS = [' - Copy', ' – Copy', '_H0', '_H1', '_C0', '_C1', ' - Copy 2', ' – Copy 2', ' - Copy 3', ' – Copy 3']
_PLACEMENT_TOKENS = ['+MU+NA+IHP+', '+MU+OFF-RS+IHP+', '+MU+OF-RH+IHP+',
                     '+IFAD+NA+IHP+', '+IFAD+NA+OSP+', '+IFAD+OF-RH+OSP+',
                     '+MU+NA+OSP+', '+IFAD-MU+NA+OSP+', '+MU+NA-NO-ID+IHP+',
                     '+MU+NA+', '+IFAD+NA+', '+FBP+', '+CATL-CAR+']

def normalize(s):
    if not s: return ''
    s = s.strip()
    for suf in _SUFFIX_TOKENS:
        while s.endswith(suf): s = s[:-len(suf)].strip()
    for pre in sorted(_PREFIX_TOKENS, key=len, reverse=True):
        if s.upper().startswith(pre.upper()):
            s = s[len(pre):]; break
    for pl in sorted(_PLACEMENT_TOKENS, key=len, reverse=True):
        s = s.replace(pl, '_')
    s = re.sub(r'[\s\+\-_\.]+', '_', s).strip('_')
    return s.upper()

def similar(a, b):
    """Return (score, method) where method describes the match strength."""
    if not a or not b: return (0.0, 'empty')
    an, bn = normalize(a), normalize(b)
    if an == bn: return (1.0, 'norm-equal')
    if len(an) >= MIN_CORE_LEN and an in bn: return (0.99, 'norm-substr')
    if len(bn) >= MIN_CORE_LEN and bn in an: return (0.99, 'norm-substr')
    ratio = difflib.SequenceMatcher(None, an, bn).ratio()
    return (ratio, f'similarity={ratio:.2f}')

# ── 1. Load source campaign ads ────────────────────────────────────
with conn.cursor() as cur:
    scope_col = 'campaign_name' if SCOPE_TYPE == 'campaign' else 'adset_name'
    cur.execute(f"""
      WITH latest AS (
        SELECT DISTINCT ON (ad_id) ad_id, ad_name, ad_status, adset_id, adset_name, account_name
          FROM public.primary_table WHERE {scope_col}=%s AND ad_id IS NOT NULL
         ORDER BY ad_id, date DESC
      ), tot AS (
        SELECT ad_id, SUM(amount_spent_inr)::numeric(14,2) AS spend
          FROM public.primary_table WHERE {scope_col}=%s GROUP BY ad_id
      )
      SELECT l.ad_id, l.ad_name, l.ad_status, l.adset_name, l.account_name,
             a.asset_id, t.spend
        FROM latest l LEFT JOIN public.ad_asset_ids a USING (ad_id) LEFT JOIN tot t USING (ad_id)
       ORDER BY t.spend DESC NULLS LAST;
    """, (CAMP, CAMP))
    src_ads = cur.fetchall()

# ── 2. Load ALL RCK ads (from rck_last30) ──────────────────────────
    cur.execute("""SELECT ad_id, ad_name, ad_status, campaign_name, adset_name,
                          asset_id, amount_spent, shopify_orders, shopify_sales
                     FROM public.rck_last30;""")
    rck_ads = cur.fetchall()
conn.close()

print(f'Source {SCOPE_TYPE}: {CAMP}   ads={len(src_ads)}')
print(f'RCK candidates : {len(rck_ads)}')

# ── 3. Find matches for each source ad ─────────────────────────────
rows = []
for s in src_ads:
    s_ad_id, s_name, s_stat, s_aset, s_acct, s_asset, s_spend = s
    matches = []
    for r in rck_ads:
        r_ad_id, r_name, r_stat, r_camp, r_aset, r_asset, r_spend, r_ord, r_sales = r
        if r_ad_id == s_ad_id: continue   # skip same-ad
        reasons = []
        # A. asset_id exact
        if s_asset and r_asset and s_asset == r_asset:
            reasons.append(f'asset={s_asset}')
        # B. name similarity
        score, method = similar(s_name, r_name)
        if score >= SIM_THRESHOLD:
            reasons.append(f'name-{method}')
        if reasons:
            matches.append((r, reasons, score))
    if matches:
        matches.sort(key=lambda m: (-m[2], -(float(m[0][6] or 0))))
        for r, reasons, score in matches:
            rows.append({
                'src_ad_id':      s_ad_id,
                'src_ad_name':    s_name,
                'src_status':     s_stat,
                'src_adset':      s_aset,
                'src_asset_id':   s_asset or '',
                'src_spend':      float(s_spend or 0),
                'match_reason':   ' | '.join(reasons),
                'similarity':     round(score, 2),
                'rck_ad_id':      r[0],
                'rck_ad_name':    r[1],
                'rck_status':     r[2],
                'rck_campaign':   r[3],
                'rck_adset':      r[4],
                'rck_asset_id':   r[5] or '',
                'rck_L30_spend':  float(r[6] or 0),
                'rck_L30_orders': int(r[7] or 0),
                'rck_L30_sales':  float(r[8] or 0),
            })
    else:
        # Emit a row anyway so the sheet documents "no match" for this source ad
        rows.append({
            'src_ad_id':      s_ad_id,
            'src_ad_name':    s_name,
            'src_status':     s_stat,
            'src_adset':      s_aset,
            'src_asset_id':   s_asset or '',
            'src_spend':      float(s_spend or 0),
            'match_reason':   '(no RCK match)',
            'similarity':     0.0,
            'rck_ad_id':      '',
            'rck_ad_name':    '',
            'rck_status':     '',
            'rck_campaign':   '',
            'rck_adset':      '',
            'rck_asset_id':   '',
            'rck_L30_spend':  0,
            'rck_L30_orders': 0,
            'rck_L30_sales':  0,
        })

# ── 4. Write to Excel ──────────────────────────────────────────────
def sanitize_sheet_name(s):
    for c in r'\/?*[]:': s = s.replace(c, '_')
    return s[:31]

OUT = _args.workbook
from openpyxl import load_workbook
if os.path.exists(OUT):
    wb = load_workbook(OUT)
    print(f'  → appending to existing workbook (had {len(wb.sheetnames)} sheets)')
else:
    wb = Workbook()
    wb.remove(wb.active)   # start clean; sheets get added below
    print(f'  → creating new workbook {OUT}')

sheet_title = sanitize_sheet_name(CAMP)
# Replace existing sheet with same title so re-runs are idempotent
if sheet_title in wb.sheetnames:
    del wb[sheet_title]
ws = wb.create_sheet(sheet_title)

COLS = ['src_ad_id', 'src_ad_name', 'src_status', 'src_adset', 'src_asset_id', 'src_spend',
        'match_reason', 'similarity',
        'rck_ad_id', 'rck_ad_name', 'rck_status', 'rck_campaign', 'rck_adset',
        'rck_asset_id', 'rck_L30_spend', 'rck_L30_orders', 'rck_L30_sales']

# Header
for c, h in enumerate(COLS, start=1):
    cell = ws.cell(row=1, column=c, value=h)
    cell.font = Font(bold=True, color='FFFFFF')
    cell.fill = PatternFill('solid', fgColor='1F4287')
    cell.alignment = Alignment(horizontal='center', vertical='center')
ws.freeze_panes = 'C2'   # freeze top row + first 2 cols

for ri, r in enumerate(rows, start=2):
    for ci, col in enumerate(COLS, start=1):
        ws.cell(row=ri, column=ci, value=r.get(col, ''))

# Number formats
for ci, col in enumerate(COLS, start=1):
    if col in ('src_spend', 'rck_L30_spend', 'rck_L30_sales'):
        for ri in range(2, len(rows) + 2):
            ws.cell(row=ri, column=ci).number_format = '"₹"#,##0'
    if col == 'similarity':
        for ri in range(2, len(rows) + 2):
            ws.cell(row=ri, column=ci).number_format = '0.00'

# Column widths
widths = {'src_ad_id':20, 'src_ad_name':50, 'src_status':14, 'src_adset':30, 'src_asset_id':14, 'src_spend':11,
          'match_reason':30, 'similarity':10,
          'rck_ad_id':20, 'rck_ad_name':50, 'rck_status':14, 'rck_campaign':45, 'rck_adset':30,
          'rck_asset_id':14, 'rck_L30_spend':13, 'rck_L30_orders':11, 'rck_L30_sales':13}
for ci, col in enumerate(COLS, start=1):
    ws.column_dimensions[get_column_letter(ci)].width = widths.get(col, 15)

# Filter
ws.auto_filter.ref = f'A1:{get_column_letter(len(COLS))}{len(rows)+1}'

# Summary tab — one row per campaign already processed, appended each run
SUM_NAME = 'Summary'
if SUM_NAME not in wb.sheetnames:
    summary = wb.create_sheet(SUM_NAME, 0)
    summary['A1'] = 'RCK Campaign Overlap — Summary'
    summary['A1'].font = Font(bold=True, size=14)
    hdr = ['sheet_name', 'scope_type', 'source_scope', 'source_ads', 'rows_in_tab',
           'ads_with_match', 'ads_without_match', 'refreshed_at']
    for c, h in enumerate(hdr, start=1):
        cell = summary.cell(row=3, column=c, value=h)
        cell.font = Font(bold=True, color='FFFFFF')
        cell.fill = PatternFill('solid', fgColor='1F4287')
    summary.column_dimensions['A'].width = 34
    summary.column_dimensions['B'].width = 46
    for c in range(3, 9):
        summary.column_dimensions[get_column_letter(c)].width = 20
    # Rules footer
    summary['A100'] = 'Rules applied:'
    summary['A100'].font = Font(bold=True)
    summary['A101'] = '  A) asset_id exact match'
    summary['A102'] = f'  B) ad_name similarity >= {SIM_THRESHOLD} (difflib) OR normalized substring (>= {MIN_CORE_LEN} chars)'
else:
    summary = wb[SUM_NAME]

# Find first empty row after the header (row 3)
import datetime as _dt
r = 4
while summary.cell(row=r, column=1).value is not None: r += 1
# Remove any prior row for this same sheet_name / scope
for check_r in range(4, r):
    if summary.cell(row=check_r, column=1).value == sheet_title:
        for cc in range(1, 9):
            summary.cell(row=check_r, column=cc).value = None
        r = check_r
        break
matched_ads_ct = len({row['src_ad_id'] for row in rows if row['match_reason'] != '(no RCK match)'})
unmatched_ads_ct = len({row['src_ad_id'] for row in rows if row['match_reason'] == '(no RCK match)'})
summary.cell(row=r, column=1, value=sheet_title)
summary.cell(row=r, column=2, value=SCOPE_TYPE)
summary.cell(row=r, column=3, value=CAMP)
summary.cell(row=r, column=4, value=len(src_ads))
summary.cell(row=r, column=5, value=len(rows))
summary.cell(row=r, column=6, value=matched_ads_ct)
summary.cell(row=r, column=7, value=unmatched_ads_ct)
summary.cell(row=r, column=8, value=_dt.datetime.now().strftime('%Y-%m-%d %H:%M:%S'))

wb.save(OUT)
print(f'\n[ok] wrote {OUT}   sheet="{ws.title}"   rows={len(rows)}')
matched_ads_ct = len({r['src_ad_id'] for r in rows if r['match_reason'] != '(no RCK match)'})
print(f'   source ads with match: {matched_ads_ct} / {len(src_ads)}')
