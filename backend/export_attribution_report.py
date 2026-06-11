"""
export_attribution_report.py - generate the Order Attribution Excel report.

Writes a multi-sheet workbook to ~/Downloads with:
  - Summary       : top-line counts, Meta vs non-Meta, match-tier breakdown
  - utm_sources   : full distribution of utm_source values
  - Top Ads       : top 50 ads by orders (T1+T2 only -- true ad_id resolved)
  - Top Campaigns : top 30 campaigns by orders (all tiers)
"""
import os, sys, io
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding="utf-8", errors="backslashreplace")

import psycopg2, openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from dotenv import load_dotenv

load_dotenv()
OUT = r"C:\Users\Saadaa\Downloads\Order_Attribution_Report.xlsx"

HEADER_FILL = PatternFill("solid", fgColor="1F4E78")
HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
SECTION_FILL = PatternFill("solid", fgColor="D9E1F2")
SECTION_FONT = Font(bold=True, size=12, color="1F4E78")
THIN = Side(border_style="thin", color="999999")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)


def style_header(ws, row, ncols):
    for c in range(1, ncols + 1):
        cell = ws.cell(row=row, column=c)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = BORDER


def autosize(ws):
    for col in ws.columns:
        m = max((len(str(c.value)) for c in col if c.value is not None), default=10)
        ws.column_dimensions[get_column_letter(col[0].column)].width = min(m + 2, 55)


def main():
    conn = psycopg2.connect(os.environ["SUPABASE_DB_URL"].strip())
    cur = conn.cursor()
    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    # ─ totals
    cur.execute("SELECT COUNT(*), ROUND(SUM(total_price),0) FROM shopify_ad_attribution")
    total_n, total_sales = cur.fetchone()
    total_sales = float(total_sales or 0)

    cur.execute("SELECT COUNT(*) FILTER (WHERE has_match) FROM shopify_ad_attribution")
    matched_total = cur.fetchone()[0]

    cur.execute("SELECT MIN(order_created_at)::date, MAX(order_created_at)::date FROM shopify_ad_attribution")
    mn, mx = cur.fetchone()

    # ──────────── Sheet 1: Summary ────────────
    ws = wb.create_sheet("Summary")
    ws["A1"] = "Saadaa | Order Attribution Report"
    ws["A1"].font = Font(bold=True, size=16, color="1F4E78")
    ws["A2"] = f"Window: {mn} -> {mx}   |   Source: shopify_ad_attribution"
    ws["A2"].font = Font(italic=True, color="666666")

    ws["A4"] = "Total orders"
    ws["B4"] = total_n
    ws["B4"].number_format = "#,##0"
    ws["A4"].font = Font(bold=True)
    ws["A5"] = "Total order value (Rs)"
    ws["B5"] = total_sales
    ws["B5"].number_format = "#,##0"
    ws["A5"].font = Font(bold=True)
    ws["A6"] = "Matched orders"
    ws["B6"] = matched_total
    ws["B6"].number_format = "#,##0"
    ws["A6"].font = Font(bold=True)
    ws["C6"] = matched_total / total_n if total_n else 0
    ws["C6"].number_format = "0.00%"

    # 1) Bucket breakdown
    ws["A8"] = "1) Orders by utm_source bucket"
    ws["A8"].font = SECTION_FONT
    ws["A8"].fill = SECTION_FILL
    hdrs = ["Bucket", "Orders", "% of total", "Sales (Rs)", "% of sales"]
    for i, h in enumerate(hdrs):
        ws.cell(row=9, column=i + 1, value=h)
    style_header(ws, 9, len(hdrs))

    cur.execute("""
        SELECT CASE WHEN UPPER(utm_source) = 'META' THEN 'Meta-tagged'
                    WHEN utm_source IS NULL OR utm_source = '' THEN 'No utm_source'
                    ELSE 'Non-Meta tagged' END AS bucket,
               COUNT(*), ROUND(SUM(total_price),0)
        FROM shopify_ad_attribution
        GROUP BY 1 ORDER BY 2 DESC
    """)
    row = 10
    for b, n, s in cur.fetchall():
        s = float(s or 0)
        ws.cell(row=row, column=1, value=b)
        ws.cell(row=row, column=2, value=n).number_format = "#,##0"
        ws.cell(row=row, column=3, value=n / total_n).number_format = "0.00%"
        ws.cell(row=row, column=4, value=s).number_format = "#,##0"
        ws.cell(row=row, column=5, value=s / total_sales if total_sales else 0).number_format = "0.00%"
        row += 1

    # 2) Match-tier breakdown
    ws.cell(row=row + 2, column=1, value="2) Match tier breakdown").font = SECTION_FONT
    ws.cell(row=row + 2, column=1).fill = SECTION_FILL
    row += 3
    hdrs = ["Match tier", "Orders", "% of all orders", "Sales (Rs)", "% of matched"]
    for i, h in enumerate(hdrs):
        ws.cell(row=row, column=i + 1, value=h)
    style_header(ws, row, len(hdrs))
    row += 1

    cur.execute("""
        SELECT COALESCE(matched_tier, '(no match)') AS tier,
               COUNT(*), ROUND(SUM(total_price), 0)
        FROM shopify_ad_attribution
        GROUP BY 1
        ORDER BY (CASE WHEN COALESCE(matched_tier,'(no match)') = '(no match)' THEN 1 ELSE 0 END),
                 2 DESC
    """)
    for tier, n, s in cur.fetchall():
        s = float(s or 0)
        ws.cell(row=row, column=1, value=tier)
        ws.cell(row=row, column=2, value=n).number_format = "#,##0"
        ws.cell(row=row, column=3, value=n / total_n).number_format = "0.00%"
        ws.cell(row=row, column=4, value=s).number_format = "#,##0"
        if tier != "(no match)":
            ws.cell(row=row, column=5, value=n / matched_total).number_format = "0.00%"
        row += 1

    autosize(ws)

    # ──────────── Sheet 2: utm_sources ────────────
    ws2 = wb.create_sheet("utm_sources")
    ws2["A1"] = "Full utm_source distribution"
    ws2["A1"].font = Font(bold=True, size=14, color="1F4E78")
    ws2["A2"] = "(every distinct utm_source value, ranked by order volume)"
    ws2["A2"].font = Font(italic=True, color="666666")
    hdrs = ["utm_source", "Orders", "% of total", "Sales (Rs)", "Avg order value (Rs)"]
    for i, h in enumerate(hdrs):
        ws2.cell(row=4, column=i + 1, value=h)
    style_header(ws2, 4, len(hdrs))

    cur.execute("""
        SELECT COALESCE(utm_source, '(NULL)'),
               COUNT(*), ROUND(SUM(total_price), 0)
        FROM shopify_ad_attribution
        GROUP BY 1 ORDER BY COUNT(*) DESC
    """)
    r = 5
    for src, n, s in cur.fetchall():
        s = float(s or 0)
        ws2.cell(row=r, column=1, value=src)
        ws2.cell(row=r, column=2, value=n).number_format = "#,##0"
        ws2.cell(row=r, column=3, value=n / total_n).number_format = "0.00%"
        ws2.cell(row=r, column=4, value=s).number_format = "#,##0"
        ws2.cell(row=r, column=5, value=s / n if n else 0).number_format = "#,##0"
        r += 1
    autosize(ws2)

    # ──────────── Sheet 3: Top Ads ────────────
    ws3 = wb.create_sheet("Top Ads")
    ws3["A1"] = "Top 50 ads by orders (T1 + T2 — true ad-level attribution)"
    ws3["A1"].font = Font(bold=True, size=14, color="1F4E78")
    hdrs = ["Rank", "ad_id", "ad_name", "campaign_name", "adset_id",
            "Orders", "Sales (Rs)", "AOV (Rs)", "First seen", "Last seen"]
    for i, h in enumerate(hdrs):
        ws3.cell(row=3, column=i + 1, value=h)
    style_header(ws3, 3, len(hdrs))

    cur.execute("""
        SELECT ad_id, ad_name, campaign_name, adset_id,
               COUNT(*), ROUND(SUM(total_price), 0),
               ROUND(AVG(total_price), 0),
               MIN(order_created_at)::date, MAX(order_created_at)::date
        FROM shopify_ad_attribution
        WHERE has_match AND matched_tier IN ('T1_ad_id','T2_ad_name_exact','T2_ad_name_fuzzy')
          AND ad_id IS NOT NULL
        GROUP BY ad_id, ad_name, campaign_name, adset_id
        ORDER BY COUNT(*) DESC LIMIT 50
    """)
    r = 4
    for ad_id, name, camp, adset, n, s, aov, fs, ls in cur.fetchall():
        s = float(s or 0); aov = float(aov or 0)
        ws3.cell(row=r, column=1, value=r - 3)
        ws3.cell(row=r, column=2, value=ad_id)
        ws3.cell(row=r, column=3, value=name)
        ws3.cell(row=r, column=4, value=camp)
        ws3.cell(row=r, column=5, value=adset)
        ws3.cell(row=r, column=6, value=n).number_format = "#,##0"
        ws3.cell(row=r, column=7, value=s).number_format = "#,##0"
        ws3.cell(row=r, column=8, value=aov).number_format = "#,##0"
        ws3.cell(row=r, column=9, value=fs)
        ws3.cell(row=r, column=10, value=ls)
        r += 1
    autosize(ws3)

    # ──────────── Sheet 4: Top Campaigns ────────────
    ws4 = wb.create_sheet("Top Campaigns")
    ws4["A1"] = "Top 30 campaigns by orders (all tiers)"
    ws4["A1"].font = Font(bold=True, size=14, color="1F4E78")
    hdrs = ["Rank", "Campaign name", "Orders", "% of matched", "Sales (Rs)", "AOV (Rs)"]
    for i, h in enumerate(hdrs):
        ws4.cell(row=3, column=i + 1, value=h)
    style_header(ws4, 3, len(hdrs))

    cur.execute("""
        SELECT campaign_name, COUNT(*), ROUND(SUM(total_price), 0), ROUND(AVG(total_price), 0)
        FROM shopify_ad_attribution
        WHERE has_match AND campaign_name IS NOT NULL AND campaign_name <> ''
        GROUP BY campaign_name ORDER BY COUNT(*) DESC LIMIT 30
    """)
    r = 4
    for camp, n, s, aov in cur.fetchall():
        s = float(s or 0); aov = float(aov or 0)
        ws4.cell(row=r, column=1, value=r - 3)
        ws4.cell(row=r, column=2, value=camp)
        ws4.cell(row=r, column=3, value=n).number_format = "#,##0"
        ws4.cell(row=r, column=4, value=n / matched_total if matched_total else 0).number_format = "0.00%"
        ws4.cell(row=r, column=5, value=s).number_format = "#,##0"
        ws4.cell(row=r, column=6, value=aov).number_format = "#,##0"
        r += 1
    autosize(ws4)

    wb.save(OUT)
    print(f"[ok] saved: {OUT}")
    print(f"  size: {os.path.getsize(OUT) / 1024:.1f} KB")
    conn.close()


if __name__ == "__main__":
    main()
